# -*- coding: utf-8 -*-
"""
Report Generator — Genera el INFORME financiero completo en Excel.

Replica la estructura del INFORME.xls profesional con:
  - EF: Estado de Situación Financiera (agregado por SUBGRUPO, comparativo)
  - ER: Estado de Resultado Integral (waterfall P&L)
  - Notas: Notas a los Estados Financieros (detalle por subcuenta)
  - Graficas: Gráficas de composición
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

# ---------------------------------------------------------------------------
# NIIF presentation names for each SUBGRUPO (used in EF)
# ---------------------------------------------------------------------------
NIIF_NAMES = {
    "11": "Efectivo y Equivalentes al Efectivo",
    "12": "Inversiones",
    "13": "Deudores comerciales y otras cuentas por cobrar",
    "14": "Inventarios",
    "15": "Propiedades, Planta y Equipo",
    "16": "Intangibles",
    "17": "Diferidos",
    "18": "Otros Activos",
    "19": "Valorizaciones",
    "21": "Obligaciones financieras",
    "22": "Proveedores",
    "23": "Cuentas comerciales por pagar y otras cuentas por pagar",
    "24": "Impuestos corrientes por pagar",
    "25": "Obligaciones laborales",
    "26": "Pasivos estimados y provisiones",
    "27": "Ingresos recibidos por anticipado",
    "28": "Otros pasivos",
    "29": "Bonos y papeles comerciales",
    "31": "Capital emitido",
    "32": "Superávit de capital",
    "33": "Reservas",
    "34": "Revalorización del patrimonio",
    "35": "Dividendos o participaciones",
    "36": "Resultado integral del año",
    "37": "Resultados de ejercicios anteriores",
    "38": "Superávit por valorizaciones",
}

ER_NIIF_NAMES = {
    "41": "Venta de producto y prestación de servicios",
    "42": "Ingresos no ordinarios",
    "47": "Ajustes por inflación (ingresos)",
    "51": "Gastos de administración",
    "52": "Gastos de Venta",
    "53": "Gastos no ordinarios",
    "54": "Provisión Impuesto de Renta",
    "59": "Ganancias y pérdidas",
    "61": "Costo de ventas y prestación de servicios",
    "62": "Compras",
    "71": "Materia prima",
    "72": "Mano de obra directa",
    "73": "Costos indirectos",
    "74": "Contratos de servicios",
}

GASTOS_4DIG_NAMES = {
    # 51 - Operacionales de administración
    "5105": "Gastos de personal", "5110": "Honorarios",
    "5115": "Impuestos", "5120": "Arrendamientos",
    "5125": "Contribuciones y afiliaciones", "5130": "Seguros",
    "5135": "Servicios", "5140": "Gastos legales",
    "5145": "Mantenimiento y reparaciones", "5150": "Adecuación e instalación",
    "5155": "Gastos de viaje", "5160": "Depreciaciones",
    "5195": "Diversos", "5199": "Provisiones",
    # 52 - Operacionales de ventas
    "5205": "Gastos de personal", "5210": "Honorarios",
    "5215": "Impuestos", "5220": "Arrendamientos",
    "5235": "Servicios", "5240": "Gastos legales",
    "5245": "Mantenimiento y reparaciones", "5295": "Diversos",
}

# Suffix C=Corriente, F=Fijo/No Corriente for the code display
SUBGRUPO_SUFFIX = {
    "11": "C", "12": "C", "13": "C", "14": "C",
    "15": "F", "16": "F", "17": "F", "18": "F", "19": "F",
    "21": "C", "22": "C", "23": "C", "24": "C", "25": "C", "26": "C", "28": "C",
    "27": "F", "29": "F",
    "31": "F", "32": "F", "33": "F", "34": "F", "35": "F", "36": "F", "37": "F", "38": "F",
}

# ---------------------------------------------------------------------------
# Colors matching the INFORME.xls teal/professional style
# ---------------------------------------------------------------------------
COL_TEAL_DARK = "1B6B6D"
COL_TEAL_MED = "3A9A9C"
COL_TEAL_LIGHT = "B5DFE0"
COL_TEAL_BG = "E0F2F2"
COL_WHITE = "FFFFFF"
COL_BLACK = "000000"
COL_DARK_TEXT = "1B1B1B"
COL_RED = "C00000"
COL_GRAY = "808080"
COL_LIGHT_GRAY = "F5F5F5"

# ---------------------------------------------------------------------------
# Reusable styles
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
BOX_BORDER = Border(
    left=Side(style="medium", color=COL_BLACK),
    right=Side(style="medium", color=COL_BLACK),
    top=Side(style="medium", color=COL_BLACK),
    bottom=Side(style="medium", color=COL_BLACK),
)

FONT_TITLE = Font(name="Arial Narrow", size=12, bold=True, color=COL_BLACK)
FONT_SUBTITLE = Font(name="Arial Narrow", size=10, bold=True, color=COL_BLACK)
FONT_SMALL_ITALIC = Font(name="Arial Narrow", size=9, italic=True, color=COL_BLACK)
FONT_HEADER = Font(name="Arial Narrow", size=9, bold=True, color=COL_BLACK)
FONT_BODY = Font(name="Arial Narrow", size=9, color=COL_BLACK)
FONT_BODY_BOLD = Font(name="Arial Narrow", size=9, bold=True, color=COL_BLACK)
FONT_TOTAL_WHITE = Font(name="Arial Narrow", size=9, bold=True, color=COL_WHITE)
FONT_SECTION = Font(name="Arial Narrow", size=9, bold=True, color=COL_BLACK)
FONT_FIRMA = Font(name="Arial Narrow", size=9, color=COL_BLACK)
FONT_FIRMA_BOLD = Font(name="Arial Narrow", size=9, bold=True, color=COL_BLACK)
FONT_FILTRO = Font(name="Arial Narrow", size=8, color=COL_GRAY)
FONT_DETAIL = Font(name="Arial Narrow", size=8, color="666666")

FILL_TEAL_DARK = PatternFill(start_color=COL_TEAL_DARK, end_color=COL_TEAL_DARK, fill_type="solid")
FILL_TEAL_MED = PatternFill(start_color=COL_TEAL_MED, end_color=COL_TEAL_MED, fill_type="solid")
FILL_TEAL_LIGHT = PatternFill(start_color=COL_TEAL_LIGHT, end_color=COL_TEAL_LIGHT, fill_type="solid")
FILL_TEAL_BG = PatternFill(start_color=COL_TEAL_BG, end_color=COL_TEAL_BG, fill_type="solid")
FILL_WHITE = PatternFill(start_color=COL_WHITE, end_color=COL_WHITE, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=COL_LIGHT_GRAY, end_color=COL_LIGHT_GRAY, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NUM_FMT = '#,##0'
NUM_FMT_NEG = '#,##0;(#,##0)'
PCT_FMT = '0.00%'


def _sc(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Style a cell."""
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format


def _wc(ws, r, c, val, **kw):
    """Write + style a cell."""
    cell = ws.cell(row=r, column=c, value=val)
    _sc(cell, **kw)
    return cell


def _money(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    return float(v)


def _sum_selected_rows(column: str, rows: list[int]) -> str:
    """Excel SUM over selected rows, avoiding hidden detail double-counting."""
    if not rows:
        return "0"
    return "=SUM(" + ",".join(f"{column}{row}" for row in rows) + ")"


def _value_variation_formula(row: int) -> str:
    """Excel formula for current minus previous period."""
    return f"=D{row}-F{row}"


def _pct_variation_formula(row: int) -> str:
    """Excel formula for variation percentage, guarded for zero previous value."""
    return f"=IF(F{row}=0,0,H{row}/ABS(F{row}))"


def _apply_variation_formulas(ws, row: int) -> None:
    """Write formulas for variation columns H and I in EF/ER rows."""
    ws.cell(row=row, column=8).value = _value_variation_formula(row)
    ws.cell(row=row, column=9).value = _pct_variation_formula(row)


def _apply_percentage_formulas(
    ws,
    rows: list[int],
    current_base_cell: str,
    previous_base_cell: str,
) -> None:
    """Write formulas for % current and % previous in EF/ER rows."""
    for row in rows:
        ws.cell(row=row, column=5).value = f"=IF({current_base_cell}=0,0,D{row}/ABS({current_base_cell}))"
        ws.cell(row=row, column=7).value = f"=IF({previous_base_cell}=0,0,F{row}/ABS({previous_base_cell}))"


# ---------------------------------------------------------------------------
# Branding
# ---------------------------------------------------------------------------
def make_branding(
    empresa: str = "",
    nit: str = "",
    representante_legal: str = "",
    representante_cc: str = "",
    contador: str = "",
    contador_tp: str = "",
    contador_cc: str = "",
    revisor_fiscal: str = "",
    revisor_tp: str = "",
    revisor_cc: str = "",
    logo_bytes: bytes | None = None,
) -> dict[str, Any]:
    return dict(
        empresa=empresa, nit=nit,
        representante_legal=representante_legal, representante_cc=representante_cc,
        contador=contador, contador_tp=contador_tp, contador_cc=contador_cc,
        revisor_fiscal=revisor_fiscal, revisor_tp=revisor_tp, revisor_cc=revisor_cc,
        logo_bytes=logo_bytes,
    )


# ---------------------------------------------------------------------------
# Aggregate data by SUBGRUPO (2-digit code)
# ---------------------------------------------------------------------------
def _aggregate_by_subgrupo(df: pd.DataFrame) -> pd.DataFrame:
    """
    Takes a df_balance or df_er (output from process_dataframe) and aggregates
    VALOR by the first 2 digits of the account code (SUBGRUPO level).
    """
    if df.empty:
        return pd.DataFrame(columns=["SUBGRUPO_COD", "SUBGRUPO", "CLASE", "GRUPO", "VALOR"])

    df = df.copy()
    # Extract 2-digit subgrupo code from the CUENTA column ("130505 - Nacionales" → "13")
    df["SUBGRUPO_COD"] = df["CUENTA"].astype(str).str.extract(r'^(\d{2})')[0]

    agg = df.groupby(["SUBGRUPO_COD", "SUBGRUPO", "CLASE", "GRUPO"], as_index=False, dropna=False).agg(
        VALOR=("VALOR", "sum")
    )
    return agg.sort_values("SUBGRUPO_COD")


def _aggregate_gastos_4dig(df: pd.DataFrame, prefix_2dig: str) -> pd.DataFrame:
    """Agrupa cuentas de gastos por prefijo de 4 dígitos."""
    if df.empty:
        return pd.DataFrame(columns=["CODE_4", "LABEL", "VALOR"])
    df = df.copy()
    df["_code"] = df["CUENTA"].astype(str).str.extract(r'^(\d+)')[0]
    mask = df["_code"].str[:2] == prefix_2dig
    df = df[mask].copy()
    if df.empty:
        return pd.DataFrame(columns=["CODE_4", "LABEL", "VALOR"])
    df["CODE_4"] = df["_code"].str[:4]
    agg = df.groupby("CODE_4", as_index=False)["VALOR"].sum()
    agg = agg[agg["VALOR"] != 0].sort_values("CODE_4")
    agg["LABEL"] = agg["CODE_4"].map(GASTOS_4DIG_NAMES).fillna("Otros gastos")
    return agg


# ---------------------------------------------------------------------------
# Helpers for drill-down detail rows
# ---------------------------------------------------------------------------
def _get_detail_for_subgrupo(df, sg_code):
    """Filter DataFrame by 2-digit subgrupo, group by CUENTA with terceros."""
    if df is None or df.empty:
        return []
    df = df.copy()
    df["_code"] = df["CUENTA"].astype(str).str.extract(r'^(\d+)')[0]
    mask = df["_code"].str[:2] == sg_code
    filtered = df[mask]
    if filtered.empty:
        return []
    result = []
    for cuenta_key, cuenta_df in filtered.groupby("CUENTA", sort=True):
        cuenta_val = cuenta_df["VALOR"].sum()
        terceros = []
        for _, t_row in cuenta_df.iterrows():
            tercero = t_row.get("TERCERO", "")
            if pd.notna(tercero) and str(tercero).strip():
                terceros.append((str(tercero), _money(t_row["VALOR"])))
        result.append((str(cuenta_key), cuenta_val, terceros))
    return result


def _write_ef_detail_rows(ws, row, detail, detail_ant_map, total_base, total_base_ant,
                          base_level=1):
    """Write cuenta + tercero detail rows with outline for EF sheet."""
    account_rows = []
    calc_rows = []
    for cuenta_key, cuenta_val, terceros in detail:
        cuenta_code = cuenta_key.split(" - ")[0].strip() if " - " in cuenta_key else ""
        cuenta_name = cuenta_key.split(" - ")[1].strip() if " - " in cuenta_key else cuenta_key

        # Get anterior value for this cuenta
        ant_info = detail_ant_map.get(cuenta_key, (0, []))
        val_ant = ant_info[0]

        # Write cuenta row
        account_row = row
        account_rows.append(account_row)
        calc_rows.append(account_row)
        _ef_data_row(ws, row, f"  {cuenta_code}", f"  {cuenta_name}", "",
                     cuenta_val, "", val_ant, "", "", "", filtro=True)
        # Override font to FONT_BODY (normal, not bold)
        for c in range(1, 10):
            ws.cell(row=row, column=c).font = FONT_BODY
        ws.row_dimensions[row].outline_level = base_level
        ws.row_dimensions[row].hidden = True
        row += 1

        # Terceros
        ant_terceros = {t[0]: t[1] for t in ant_info[1]} if len(ant_info) > 1 else {}
        detail_start_row = row
        for tercero_name, tercero_val in terceros:
            t_val_ant = ant_terceros.get(tercero_name, 0)

            _ef_data_row(ws, row, "", f"    {tercero_name}", "",
                         tercero_val, "", t_val_ant, "", "", "", filtro=True)
            _apply_variation_formulas(ws, row)
            calc_rows.append(row)
            for c in range(1, 10):
                ws.cell(row=row, column=c).font = FONT_DETAIL
            ws.row_dimensions[row].outline_level = base_level + 1
            ws.row_dimensions[row].hidden = True
            row += 1

        detail_end_row = row - 1
        if detail_start_row <= detail_end_row:
            ws.cell(row=account_row, column=4).value = f"=SUM(D{detail_start_row}:D{detail_end_row})"
            ws.cell(row=account_row, column=6).value = f"=SUM(F{detail_start_row}:F{detail_end_row})"
        _apply_variation_formulas(ws, account_row)

    return row, account_rows, calc_rows


def _build_detail_ant_map(df_ant, sg_code):
    """Build a dict {CUENTA: (total_val, [(tercero, val), ...])} for anterior."""
    if df_ant is None or df_ant.empty:
        return {}
    df_ant = df_ant.copy()
    df_ant["_code"] = df_ant["CUENTA"].astype(str).str.extract(r'^(\d+)')[0]
    mask = df_ant["_code"].str[:2] == sg_code
    filtered = df_ant[mask]
    if filtered.empty:
        return {}
    result = {}
    for cuenta_key, cuenta_df in filtered.groupby("CUENTA", sort=True):
        cuenta_val = cuenta_df["VALOR"].sum()
        terceros = []
        for _, t_row in cuenta_df.iterrows():
            tercero = t_row.get("TERCERO", "")
            if pd.notna(tercero) and str(tercero).strip():
                terceros.append((str(tercero), _money(t_row["VALOR"])))
        result[str(cuenta_key)] = (cuenta_val, terceros)
    return result


def _write_er_detail_rows(ws, row, df_er, df_er_ant, sg_code, total_ingresos, total_ingresos_ant,
                          base_level=1):
    """Write cuenta + tercero detail rows for an ER subgrupo."""
    detail = _get_detail_for_subgrupo(df_er, sg_code)
    detail_ant_map = _build_detail_ant_map(df_er_ant, sg_code)
    account_rows = []
    calc_rows = []

    for cuenta_key, cuenta_val, terceros in detail:
        cuenta_code = cuenta_key.split(" - ")[0].strip() if " - " in cuenta_key else ""
        cuenta_name = cuenta_key.split(" - ")[1].strip() if " - " in cuenta_key else cuenta_key

        ant_info = detail_ant_map.get(cuenta_key, (0, []))
        val_ant = ant_info[0]

        indent = "  " * base_level
        account_row = row
        account_rows.append(account_row)
        calc_rows.append(account_row)
        c = 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER); c += 1
        _wc(ws, row, c, f"{indent}{cuenta_code} - {cuenta_name}", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
        _wc(ws, row, c, cuenta_val, font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, val_ant, font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, "TRUE", font=FONT_FILTRO)
        ws.row_dimensions[row].outline_level = base_level
        ws.row_dimensions[row].hidden = True
        row += 1

        # Terceros
        ant_terceros = {t[0]: t[1] for t in ant_info[1]} if len(ant_info) > 1 else {}
        detail_start_row = row
        for tercero_name, tercero_val in terceros:
            t_val_ant = ant_terceros.get(tercero_name, 0)

            indent2 = "  " * (base_level + 1)
            c = 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER); c += 1
            _wc(ws, row, c, f"{indent2}{tercero_name}", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
            _wc(ws, row, c, tercero_val, font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, t_val_ant, font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "TRUE", font=FONT_FILTRO)
            _apply_variation_formulas(ws, row)
            calc_rows.append(row)
            ws.row_dimensions[row].outline_level = base_level + 1
            ws.row_dimensions[row].hidden = True
            row += 1

        detail_end_row = row - 1
        if detail_start_row <= detail_end_row:
            ws.cell(row=account_row, column=4).value = f"=SUM(D{detail_start_row}:D{detail_end_row})"
            ws.cell(row=account_row, column=6).value = f"=SUM(F{detail_start_row}:F{detail_end_row})"
        _apply_variation_formulas(ws, account_row)

    return row, account_rows, calc_rows


def _get_detail_for_4dig(df, code4):
    """Filter DataFrame by 4-digit prefix, group by CUENTA with terceros."""
    if df is None or df.empty:
        return []
    df = df.copy()
    df["_code"] = df["CUENTA"].astype(str).str.extract(r'^(\d+)')[0]
    mask = df["_code"].str[:4] == code4
    filtered = df[mask]
    if filtered.empty:
        return []
    result = []
    for cuenta_key, cuenta_df in filtered.groupby("CUENTA", sort=True):
        cuenta_val = cuenta_df["VALOR"].sum()
        terceros = []
        for _, t_row in cuenta_df.iterrows():
            tercero = t_row.get("TERCERO", "")
            if pd.notna(tercero) and str(tercero).strip():
                terceros.append((str(tercero), _money(t_row["VALOR"])))
        result.append((str(cuenta_key), cuenta_val, terceros))
    return result


def _build_detail_ant_map_4dig(df_ant, code4):
    """Build anterior map for a 4-digit prefix."""
    if df_ant is None or df_ant.empty:
        return {}
    df_ant = df_ant.copy()
    df_ant["_code"] = df_ant["CUENTA"].astype(str).str.extract(r'^(\d+)')[0]
    mask = df_ant["_code"].str[:4] == code4
    filtered = df_ant[mask]
    if filtered.empty:
        return {}
    result = {}
    for cuenta_key, cuenta_df in filtered.groupby("CUENTA", sort=True):
        cuenta_val = cuenta_df["VALOR"].sum()
        terceros = []
        for _, t_row in cuenta_df.iterrows():
            tercero = t_row.get("TERCERO", "")
            if pd.notna(tercero) and str(tercero).strip():
                terceros.append((str(tercero), _money(t_row["VALOR"])))
        result[str(cuenta_key)] = (cuenta_val, terceros)
    return result


def _write_er_4dig_detail_rows(ws, row, df_er, df_er_ant, code4,
                                total_ingresos, total_ingresos_ant, base_level=2):
    """Write cuenta + tercero detail rows for a 4-digit gastos category."""
    detail = _get_detail_for_4dig(df_er, code4)
    detail_ant_map = _build_detail_ant_map_4dig(df_er_ant, code4)
    account_rows = []
    calc_rows = []

    for cuenta_key, cuenta_val, terceros in detail:
        cuenta_code = cuenta_key.split(" - ")[0].strip() if " - " in cuenta_key else ""
        cuenta_name = cuenta_key.split(" - ")[1].strip() if " - " in cuenta_key else cuenta_key

        ant_info = detail_ant_map.get(cuenta_key, (0, []))
        val_ant = ant_info[0]

        indent = "  " * base_level
        account_row = row
        account_rows.append(account_row)
        calc_rows.append(account_row)
        c = 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER); c += 1
        _wc(ws, row, c, f"{indent}{cuenta_code} - {cuenta_name}", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
        _wc(ws, row, c, cuenta_val, font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, val_ant, font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
        _wc(ws, row, c, "", font=FONT_BODY, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
        _wc(ws, row, c, "TRUE", font=FONT_FILTRO)
        ws.row_dimensions[row].outline_level = base_level
        ws.row_dimensions[row].hidden = True
        row += 1

        # Terceros
        ant_terceros = {t[0]: t[1] for t in ant_info[1]} if len(ant_info) > 1 else {}
        detail_start_row = row
        for tercero_name, tercero_val in terceros:
            t_val_ant = ant_terceros.get(tercero_name, 0)

            indent2 = "  " * (base_level + 1)
            c = 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER); c += 1
            _wc(ws, row, c, f"{indent2}{tercero_name}", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
            _wc(ws, row, c, tercero_val, font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, t_val_ant, font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=FONT_DETAIL, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "TRUE", font=FONT_FILTRO)
            _apply_variation_formulas(ws, row)
            calc_rows.append(row)
            ws.row_dimensions[row].outline_level = base_level + 1
            ws.row_dimensions[row].hidden = True
            row += 1

        detail_end_row = row - 1
        if detail_start_row <= detail_end_row:
            ws.cell(row=account_row, column=4).value = f"=SUM(D{detail_start_row}:D{detail_end_row})"
            ws.cell(row=account_row, column=6).value = f"=SUM(F{detail_start_row}:F{detail_end_row})"
        _apply_variation_formulas(ws, account_row)

    return row, account_rows, calc_rows


# ---------------------------------------------------------------------------
# Header block
# ---------------------------------------------------------------------------
def _write_header(ws, branding, title, subtitle, start_row=1, merge_cols=8):
    row = start_row

    if branding.get("logo_bytes"):
        try:
            img = XlImage(BytesIO(branding["logo_bytes"]))
            img.width, img.height = 150, 50
            ws.add_image(img, f"A{row}")
        except Exception:
            pass
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    _wc(ws, row, 1, branding.get("empresa", ""), font=FONT_TITLE, alignment=ALIGN_CENTER)
    row += 1

    if branding.get("nit"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
        _wc(ws, row, 1, f"NIT. {branding['nit']}", font=FONT_SUBTITLE, alignment=ALIGN_CENTER)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    _wc(ws, row, 1, title, font=FONT_SUBTITLE, alignment=ALIGN_CENTER)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_cols)
    _wc(ws, row, 1, subtitle, font=FONT_SMALL_ITALIC, alignment=ALIGN_CENTER)
    row += 2

    return row


# ---------------------------------------------------------------------------
# Signature block
# ---------------------------------------------------------------------------
def _write_signature_line(ws, row, start_col, end_col, value, font):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    _wc(ws, row, start_col, value, font=font, alignment=ALIGN_CENTER)


def _write_signature_block(ws, row, start_col, end_col, name, role, extra_lines):
    _write_signature_line(ws, row, start_col, end_col, name, FONT_FIRMA_BOLD)
    row += 1
    _write_signature_line(ws, row, start_col, end_col, role, FONT_FIRMA)
    row += 1

    for line in extra_lines:
        if not line:
            continue
        _write_signature_line(ws, row, start_col, end_col, line, FONT_FIRMA)
        row += 1


def _write_signatures(ws, branding, start_row, col_left=1, col_right=6):
    row = start_row + 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    _wc(ws, row, 1, "Las notas adjuntas son parte integral de este estado financiero.",
        font=Font(name="Arial Narrow", size=8, italic=True, color=COL_GRAY), alignment=ALIGN_CENTER)
    row += 3

    has_revisor = any(
        branding.get(key) for key in ("revisor_fiscal", "revisor_tp", "revisor_cc")
    )
    signature_columns = [(1, 3), (6, 8)]
    if has_revisor:
        signature_columns = [(1, 3), (4, 6), (7, 8)]

    rep_cols, contador_cols = signature_columns[:2]
    _write_signature_block(
        ws,
        row,
        rep_cols[0],
        rep_cols[1],
        branding.get("representante_legal", ""),
        "Representante Legal",
        [f"C.C. {branding['representante_cc']}" if branding.get("representante_cc") else ""],
    )
    _write_signature_block(
        ws,
        row,
        contador_cols[0],
        contador_cols[1],
        branding.get("contador", ""),
        "Contadora Pública" if branding.get("contador") else "",
        [
            f"T.P. {branding['contador_tp']}" if branding.get("contador_tp") else "",
            f"C.C {branding['contador_cc']}" if branding.get("contador_cc") else "",
        ],
    )

    if has_revisor:
        revisor_cols = signature_columns[2]
        _write_signature_block(
            ws,
            row,
            revisor_cols[0],
            revisor_cols[1],
            branding.get("revisor_fiscal", ""),
            "Revisor Fiscal",
            [
                f"T.P. {branding['revisor_tp']}" if branding.get("revisor_tp") else "",
                f"C.C {branding['revisor_cc']}" if branding.get("revisor_cc") else "",
            ],
        )


# ---------------------------------------------------------------------------
# Write a data row in the EF/ER format
# ---------------------------------------------------------------------------
def _ef_data_row(ws, row, code, name, nota, val_actual, pct_actual, val_ant, pct_ant, var_val, var_pct,
                 is_total=False, is_grand_total=False, is_section_header=False, filtro=True,
                 col_start=1):
    """Write one row in the EF or ER layout."""

    if is_grand_total:
        font = FONT_TOTAL_WHITE
        fill = FILL_TEAL_DARK
    elif is_total:
        font = FONT_BODY_BOLD
        fill = FILL_TEAL_LIGHT
    elif is_section_header:
        font = FONT_SECTION
        fill = FILL_WHITE
    else:
        font = FONT_BODY
        fill = FILL_WHITE

    neg_font_actual = Font(name="Arial Narrow", size=9, bold=is_total or is_grand_total,
                            color=COL_WHITE if is_grand_total else COL_BLACK)

    c = col_start
    # Col A: subgrupo code
    _wc(ws, row, c, code if code else "", font=font, fill=fill, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
    # Col B: name
    _wc(ws, row, c, name, font=font, fill=fill, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
    # Col C: nota
    _wc(ws, row, c, nota if nota else "", font=font, fill=fill, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
    # Col D: valor actual
    _wc(ws, row, c, val_actual, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
    # Col E: % actual
    _wc(ws, row, c, pct_actual, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
    # Col F: valor anterior
    _wc(ws, row, c, val_ant, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
    # Col G: % anterior
    _wc(ws, row, c, pct_ant, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
    # Col H: variación
    _wc(ws, row, c, var_val, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
    # Col I: % variación
    _wc(ws, row, c, var_pct, font=neg_font_actual, fill=fill, border=THIN_BORDER,
        alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
    # Col J: FILTRO
    _wc(ws, row, c, "TRUE" if filtro else "FALSE", font=FONT_FILTRO, alignment=ALIGN_CENTER)


# ---------------------------------------------------------------------------
# EF — Estado de Situación Financiera
# ---------------------------------------------------------------------------
def _build_ef_sheet(wb, df_balance, branding, periodo_actual, periodo_anterior=None, df_balance_anterior=None):
    ws = wb.create_sheet("EF")
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False)
    merge_cols = 9

    row = _write_header(ws, branding, "ESTADO DE SITUACIÓN FINANCIERA",
                        "(Expresados en pesos colombianos)", merge_cols=merge_cols)

    # "Período terminado a"
    _wc(ws, row, 1, "Período terminado a", font=FONT_SMALL_ITALIC); row += 1

    # Column headers
    per_ant_label = periodo_anterior or "Período Anterior"
    headers = ["", "ACTIVO", "Nota", periodo_actual, "%", per_ant_label, "%", "Variación", "%", "FILTRO"]
    for i, h in enumerate(headers, 1):
        _wc(ws, row, i, h, font=FONT_HEADER, fill=FILL_TEAL_BG, border=THIN_BORDER, alignment=ALIGN_CENTER)
    row += 1

    agg = _aggregate_by_subgrupo(df_balance)
    agg_ant = _aggregate_by_subgrupo(df_balance_anterior) if df_balance_anterior is not None else None

    total_activo = agg.loc[agg["CLASE"] == "ACTIVO", "VALOR"].sum()
    if total_activo == 0:
        total_activo = 1

    total_activo_ant = 0
    if agg_ant is not None:
        total_activo_ant = agg_ant.loc[agg_ant["CLASE"] == "ACTIVO", "VALOR"].sum()

    nota_num = 1

    sections_ef = [
        ("ACTIVO", [
            ("CORRIENTE", "ACTIVO CORRIENTE",
             ["11", "12", "13", "14"]),
            ("NO CORRIENTE", "ACTIVO NO CORRIENTE",
             ["15", "16", "17", "18", "19"]),
        ]),
        ("PASIVO", [
            ("PASIVO CORRIENTE", "PASIVO CORRIENTE",
             ["21", "22", "23", "24", "25", "26", "28"]),
            ("PASIVO NO CORRIENTE", "PASIVO NO CORRIENTE",
             ["27", "29"]),
        ]),
        ("PATRIMONIO", [
            ("PATRIMONIO", "PATRIMONIO",
             ["31", "32", "33", "34", "35", "36", "37", "38"]),
        ]),
    ]

    grand_totals = {}
    calc_rows = []
    active_total_row = None

    for clase_label, groups in sections_ef:
        if clase_label == "PASIVO":
            row += 1  # blank before PASIVO

        clase_total_rows = []

        for section_title, grupo_name, subgrupo_codes in groups:
            # Section header (e.g., "CORRIENTE", "NO CORRIENTE")
            _ef_data_row(ws, row, "", section_title, "", "", "", "", "", "", "",
                         is_section_header=True, filtro=True)
            row += 1

            section_rows = []

            for sg_code in subgrupo_codes:
                sg_data = agg[agg["SUBGRUPO_COD"] == sg_code]
                val = sg_data["VALOR"].sum() if not sg_data.empty else 0

                val_ant = 0
                if agg_ant is not None:
                    sg_ant_data = agg_ant[agg_ant["SUBGRUPO_COD"] == sg_code]
                    val_ant = sg_ant_data["VALOR"].sum() if not sg_ant_data.empty else 0

                if val == 0 and val_ant == 0:
                    continue

                niif_name = NIIF_NAMES.get(sg_code, sg_code)
                suffix = SUBGRUPO_SUFFIX.get(sg_code, "")
                display_code = f"{sg_code} {sg_code}{suffix}"

                note_val = nota_num
                nota_num += 1

                subgrupo_row = row
                section_rows.append(subgrupo_row)
                calc_rows.append(subgrupo_row)
                _ef_data_row(ws, row, display_code, niif_name, note_val,
                             val, "", val_ant, "", "", "", filtro=True)
                row += 1

                # Detail: cuentas + terceros (collapsed)
                detail = _get_detail_for_subgrupo(df_balance, sg_code)
                detail_ant_map = _build_detail_ant_map(df_balance_anterior, sg_code)
                row, detail_account_rows, detail_calc_rows = _write_ef_detail_rows(
                    ws, row, detail, detail_ant_map,
                    total_activo, total_activo_ant, base_level=1
                )
                calc_rows.extend(detail_calc_rows)
                if detail_account_rows:
                    ws.cell(row=subgrupo_row, column=4).value = _sum_selected_rows("D", detail_account_rows)
                    ws.cell(row=subgrupo_row, column=6).value = _sum_selected_rows("F", detail_account_rows)
                _apply_variation_formulas(ws, subgrupo_row)

            # Section total
            section_total_row = row
            clase_total_rows.append(section_total_row)
            calc_rows.append(section_total_row)
            _ef_data_row(ws, row, "", f"TOTAL {section_title}", "",
                         _sum_selected_rows("D", section_rows), "",
                         _sum_selected_rows("F", section_rows), "", "", "",
                         is_total=True, filtro=True)
            _apply_variation_formulas(ws, row)
            row += 1

        # Grand total for clase (TOTAL ACTIVO, TOTAL PASIVO, TOTAL PATRIMONIO)
        clase_total_row = row
        calc_rows.append(clase_total_row)
        _ef_data_row(ws, row, "", f"TOTAL {clase_label}", "",
                     _sum_selected_rows("D", clase_total_rows), "",
                     _sum_selected_rows("F", clase_total_rows), "", "", "",
                     is_grand_total=True, filtro=True)
        _apply_variation_formulas(ws, row)
        if clase_label == "ACTIVO":
            active_total_row = clase_total_row
        row += 1

        grand_totals[clase_label] = clase_total_row

    # TOTAL PASIVO Y PATRIMONIO
    row += 1
    pasivo_patrimonio_rows = [
        total_row for total_row in [
            grand_totals.get("PASIVO"),
            grand_totals.get("PATRIMONIO"),
        ]
        if total_row
    ]
    calc_rows.append(row)
    _ef_data_row(ws, row, "", "TOTAL PASIVO Y PATRIMONIO", "",
                 _sum_selected_rows("D", pasivo_patrimonio_rows), "",
                 _sum_selected_rows("F", pasivo_patrimonio_rows), "", "", "",
                 is_grand_total=True, filtro=True)
    _apply_variation_formulas(ws, row)
    row += 2

    if active_total_row:
        _apply_percentage_formulas(
            ws,
            calc_rows,
            current_base_cell=f"$D${active_total_row}",
            previous_base_cell=f"$F${active_total_row}",
        )

    _write_signatures(ws, branding, row)

    # Column widths
    widths = {"A": 10, "B": 52, "C": 6, "D": 16, "E": 9, "F": 16, "G": 9, "H": 16, "I": 9, "J": 8}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    return nota_num


# ---------------------------------------------------------------------------
# ER — Estado de Resultado Integral (P&L waterfall)
# ---------------------------------------------------------------------------
def _build_er_sheet(wb, df_er, branding, periodo_actual, nota_start=1,
                    periodo_anterior=None, df_er_anterior=None):
    ws = wb.create_sheet("ER")
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False)
    merge_cols = 9

    row = _write_header(ws, branding, "ESTADO DE RESULTADO INTEGRAL",
                        "(Expresados en pesos colombianos)", merge_cols=merge_cols)

    # Period label
    _wc(ws, row, 1, "Período comprendido entre", font=FONT_SMALL_ITALIC); row += 1
    _wc(ws, row, 1, f"01 de Enero y al fin del mes de : {periodo_actual}", font=FONT_SMALL_ITALIC); row += 1
    row += 1

    # Column headers
    per_ant_label = periodo_anterior or "Período Anterior"
    headers = ["Actividades Ordinarias", "", "Nota", periodo_actual, "%", per_ant_label, "%", "VARIACIÓN", "%", "FILTRO"]
    for i, h in enumerate(headers, 1):
        _wc(ws, row, i, h, font=FONT_HEADER, fill=FILL_TEAL_BG, border=THIN_BORDER, alignment=ALIGN_CENTER)
    row += 1

    agg = _aggregate_by_subgrupo(df_er)
    agg_ant = _aggregate_by_subgrupo(df_er_anterior) if df_er_anterior is not None else None

    def _get_val(sg_code):
        d = agg[agg["SUBGRUPO_COD"] == sg_code]
        return d["VALOR"].sum() if not d.empty else 0

    def _get_val_ant(sg_code):
        if agg_ant is None:
            return 0
        d = agg_ant[agg_ant["SUBGRUPO_COD"] == sg_code]
        return d["VALOR"].sum() if not d.empty else 0

    # Ingresos operacionales (41)
    ingresos_41 = _get_val("41")
    ingresos_41_ant = _get_val_ant("41")
    total_ingresos = abs(ingresos_41) if ingresos_41 != 0 else 1
    total_ingresos_ant = abs(ingresos_41_ant) if ingresos_41_ant != 0 else 1

    def _pct(val, base):
        return val / base if base != 0 else 0

    def _var_pct(val, val_ant):
        return (val - val_ant) / abs(val_ant) if val_ant != 0 else 0

    nota = nota_start
    calc_rows = []
    ingresos_row = None

    def _row_item(label, val, val_ant, note=None, bold=False, grand=False, total_bg=False):
        nonlocal row, nota
        written_row = row
        n = None
        if note:
            n = nota
            nota += 1

        if grand:
            _ef_data_row(ws, row, "", label, n, val, "",
                         val_ant, "", "", "", is_grand_total=True)
        elif total_bg:
            _ef_data_row(ws, row, "", label, n, val, "",
                         val_ant, "", "", "", is_total=True)
        else:
            font = FONT_BODY_BOLD if bold else FONT_BODY
            c = 1
            _wc(ws, row, c, "", font=font, fill=FILL_WHITE, border=THIN_BORDER); c += 1
            _wc(ws, row, c, label, font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_LEFT); c += 1
            _wc(ws, row, c, n if n else "", font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_CENTER); c += 1
            _wc(ws, row, c, val, font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, val_ant, font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "", font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG); c += 1
            _wc(ws, row, c, "", font=font, fill=FILL_WHITE, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT); c += 1
            _wc(ws, row, c, "TRUE", font=FONT_FILTRO)

        _apply_variation_formulas(ws, written_row)
        row += 1
        calc_rows.append(written_row)
        return written_row

    # --- P&L Waterfall ---

    # Ventas (41)
    ingresos_row = _row_item("Venta de producto y prestación de servicios", ingresos_41, ingresos_41_ant, note=True)
    # Detail for 41 (level 1 = cuenta, level 2 = tercero)
    row, ingresos_account_rows, ingresos_calc_rows = _write_er_detail_rows(
        ws, row, df_er, df_er_anterior, "41",
        total_ingresos, total_ingresos_ant, base_level=1
    )
    calc_rows.extend(ingresos_calc_rows)
    if ingresos_account_rows:
        ws.cell(row=ingresos_row, column=4).value = _sum_selected_rows("D", ingresos_account_rows)
        ws.cell(row=ingresos_row, column=6).value = _sum_selected_rows("F", ingresos_account_rows)
        _apply_variation_formulas(ws, ingresos_row)

    # Costos de ventas (61, 62, 71-74)
    costo_codes = ["61", "62", "71", "72", "73", "74"]
    costos = sum(_get_val(c) for c in costo_codes)
    costos_ant = sum(_get_val_ant(c) for c in costo_codes)
    costos_row = _row_item("MENOS:  COSTO DE VENTAS", costos, costos_ant)
    # Detail for each costo code
    costo_account_rows = []
    for costo_sg in costo_codes:
        row, account_rows, detail_calc_rows = _write_er_detail_rows(
            ws, row, df_er, df_er_anterior, costo_sg,
            total_ingresos, total_ingresos_ant, base_level=1
        )
        costo_account_rows.extend(account_rows)
        calc_rows.extend(detail_calc_rows)
    if costo_account_rows:
        ws.cell(row=costos_row, column=4).value = _sum_selected_rows("D", costo_account_rows)
        ws.cell(row=costos_row, column=6).value = _sum_selected_rows("F", costo_account_rows)
        _apply_variation_formulas(ws, costos_row)

    # UTILIDAD BRUTA
    util_bruta = ingresos_41 - costos
    util_bruta_ant = ingresos_41_ant - costos_ant
    util_bruta_row = _row_item("UTILIDAD BRUTA", util_bruta, util_bruta_ant, total_bg=True)
    ws.cell(row=util_bruta_row, column=4).value = f"=D{ingresos_row}-D{costos_row}"
    ws.cell(row=util_bruta_row, column=6).value = f"=F{ingresos_row}-F{costos_row}"
    _apply_variation_formulas(ws, util_bruta_row)

    # --- Gastos de administración (51) - desglosado ---
    gastos_admin = _get_val("51")
    gastos_admin_ant = _get_val_ant("51")

    sub_51 = _aggregate_gastos_4dig(df_er, "51")
    sub_51_ant = _aggregate_gastos_4dig(df_er_anterior, "51") if df_er_anterior is not None else pd.DataFrame()

    all_codes_51 = sorted(set(
        sub_51["CODE_4"].tolist() +
        (sub_51_ant["CODE_4"].tolist() if not sub_51_ant.empty else [])
    ))

    admin_rows = []
    for code4 in all_codes_51:
        label = GASTOS_4DIG_NAMES.get(code4, "Otros gastos")
        val = sub_51.loc[sub_51["CODE_4"] == code4, "VALOR"].sum() if not sub_51.empty else 0
        val_ant = sub_51_ant.loc[sub_51_ant["CODE_4"] == code4, "VALOR"].sum() if not sub_51_ant.empty else 0
        if val == 0 and val_ant == 0:
            continue
        code4_row = _row_item(f"  {label}", val, val_ant)
        admin_rows.append(code4_row)
        # Mark 4-digit row as outline_level=1
        ws.row_dimensions[row - 1].outline_level = 1
        ws.row_dimensions[row - 1].hidden = True
        # Detail: cuentas + terceros within this 4-digit category
        row, account_rows, detail_calc_rows = _write_er_4dig_detail_rows(
            ws, row, df_er, df_er_anterior, code4,
            total_ingresos, total_ingresos_ant, base_level=2
        )
        if account_rows:
            ws.cell(row=code4_row, column=4).value = _sum_selected_rows("D", account_rows)
            ws.cell(row=code4_row, column=6).value = _sum_selected_rows("F", account_rows)
            _apply_variation_formulas(ws, code4_row)
        calc_rows.extend(detail_calc_rows)

    gastos_admin_row = _row_item("Total Gastos de administración", gastos_admin, gastos_admin_ant, note=True)
    if admin_rows:
        ws.cell(row=gastos_admin_row, column=4).value = _sum_selected_rows("D", admin_rows)
        ws.cell(row=gastos_admin_row, column=6).value = _sum_selected_rows("F", admin_rows)
        _apply_variation_formulas(ws, gastos_admin_row)

    # --- Gastos de venta (52) - desglosado ---
    gastos_venta = _get_val("52")
    gastos_venta_ant = _get_val_ant("52")

    sub_52 = _aggregate_gastos_4dig(df_er, "52")
    sub_52_ant = _aggregate_gastos_4dig(df_er_anterior, "52") if df_er_anterior is not None else pd.DataFrame()

    all_codes_52 = sorted(set(
        sub_52["CODE_4"].tolist() +
        (sub_52_ant["CODE_4"].tolist() if not sub_52_ant.empty else [])
    ))

    venta_rows = []
    for code4 in all_codes_52:
        label = GASTOS_4DIG_NAMES.get(code4, "Otros gastos")
        val = sub_52.loc[sub_52["CODE_4"] == code4, "VALOR"].sum() if not sub_52.empty else 0
        val_ant = sub_52_ant.loc[sub_52_ant["CODE_4"] == code4, "VALOR"].sum() if not sub_52_ant.empty else 0
        if val == 0 and val_ant == 0:
            continue
        code4_row = _row_item(f"  {label}", val, val_ant)
        venta_rows.append(code4_row)
        # Mark 4-digit row as outline_level=1
        ws.row_dimensions[row - 1].outline_level = 1
        ws.row_dimensions[row - 1].hidden = True
        # Detail: cuentas + terceros within this 4-digit category
        row, account_rows, detail_calc_rows = _write_er_4dig_detail_rows(
            ws, row, df_er, df_er_anterior, code4,
            total_ingresos, total_ingresos_ant, base_level=2
        )
        if account_rows:
            ws.cell(row=code4_row, column=4).value = _sum_selected_rows("D", account_rows)
            ws.cell(row=code4_row, column=6).value = _sum_selected_rows("F", account_rows)
            _apply_variation_formulas(ws, code4_row)
        calc_rows.extend(detail_calc_rows)

    gastos_venta_row = _row_item("Total Gastos de Venta", gastos_venta, gastos_venta_ant)
    if venta_rows:
        ws.cell(row=gastos_venta_row, column=4).value = _sum_selected_rows("D", venta_rows)
        ws.cell(row=gastos_venta_row, column=6).value = _sum_selected_rows("F", venta_rows)
        _apply_variation_formulas(ws, gastos_venta_row)

    # UTILIDAD OPERACIONAL
    util_oper = util_bruta - gastos_admin - gastos_venta
    util_oper_ant = util_bruta_ant - gastos_admin_ant - gastos_venta_ant
    util_oper_row = _row_item("UTILIDAD OPERACIONAL", util_oper, util_oper_ant, grand=True)
    ws.cell(row=util_oper_row, column=4).value = f"=D{util_bruta_row}-D{gastos_admin_row}-D{gastos_venta_row}"
    ws.cell(row=util_oper_row, column=6).value = f"=F{util_bruta_row}-F{gastos_admin_row}-F{gastos_venta_row}"
    _apply_variation_formulas(ws, util_oper_row)

    # Ingresos no ordinarios (42)
    ing_no_ord = _get_val("42")
    ing_no_ord_ant = _get_val_ant("42")
    ing_no_ord_row = _row_item("Ingresos no ordinarios", ing_no_ord, ing_no_ord_ant, note=True)
    row, account_rows, detail_calc_rows = _write_er_detail_rows(
        ws, row, df_er, df_er_anterior, "42",
        total_ingresos, total_ingresos_ant, base_level=1
    )
    if account_rows:
        ws.cell(row=ing_no_ord_row, column=4).value = _sum_selected_rows("D", account_rows)
        ws.cell(row=ing_no_ord_row, column=6).value = _sum_selected_rows("F", account_rows)
        _apply_variation_formulas(ws, ing_no_ord_row)
    calc_rows.extend(detail_calc_rows)

    # Gastos no ordinarios (53)
    gtos_no_ord = _get_val("53")
    gtos_no_ord_ant = _get_val_ant("53")
    gtos_no_ord_row = _row_item("Gastos no ordinarios", gtos_no_ord, gtos_no_ord_ant, note=True)
    row, account_rows, detail_calc_rows = _write_er_detail_rows(
        ws, row, df_er, df_er_anterior, "53",
        total_ingresos, total_ingresos_ant, base_level=1
    )
    if account_rows:
        ws.cell(row=gtos_no_ord_row, column=4).value = _sum_selected_rows("D", account_rows)
        ws.cell(row=gtos_no_ord_row, column=6).value = _sum_selected_rows("F", account_rows)
        _apply_variation_formulas(ws, gtos_no_ord_row)
    calc_rows.extend(detail_calc_rows)

    # UTILIDAD (PERDIDA) ANTES DE IMPUESTOS
    util_antes_imp = util_oper + ing_no_ord - gtos_no_ord
    util_antes_imp_ant = util_oper_ant + ing_no_ord_ant - gtos_no_ord_ant
    util_antes_imp_row = _row_item("UTILIDAD (PERDIDA) ANTES DE IMPUESTOS", util_antes_imp, util_antes_imp_ant, grand=True)
    ws.cell(row=util_antes_imp_row, column=4).value = f"=D{util_oper_row}+D{ing_no_ord_row}-D{gtos_no_ord_row}"
    ws.cell(row=util_antes_imp_row, column=6).value = f"=F{util_oper_row}+F{ing_no_ord_row}-F{gtos_no_ord_row}"
    _apply_variation_formulas(ws, util_antes_imp_row)

    # Provisión Impuesto de Renta (54)
    imp_renta = _get_val("54")
    imp_renta_ant = _get_val_ant("54")
    imp_renta_row = _row_item("Provisión Impuesto de Renta", imp_renta, imp_renta_ant)
    row, account_rows, detail_calc_rows = _write_er_detail_rows(
        ws, row, df_er, df_er_anterior, "54",
        total_ingresos, total_ingresos_ant, base_level=1
    )
    if account_rows:
        ws.cell(row=imp_renta_row, column=4).value = _sum_selected_rows("D", account_rows)
        ws.cell(row=imp_renta_row, column=6).value = _sum_selected_rows("F", account_rows)
        _apply_variation_formulas(ws, imp_renta_row)
    calc_rows.extend(detail_calc_rows)

    # UTILIDAD NETA
    util_neta = util_antes_imp - imp_renta
    util_neta_ant = util_antes_imp_ant - imp_renta_ant
    util_neta_row = _row_item("UTILIDAD NETA", util_neta, util_neta_ant, grand=True)
    ws.cell(row=util_neta_row, column=4).value = f"=D{util_antes_imp_row}-D{imp_renta_row}"
    ws.cell(row=util_neta_row, column=6).value = f"=F{util_antes_imp_row}-F{imp_renta_row}"
    _apply_variation_formulas(ws, util_neta_row)

    if ingresos_row:
        _apply_percentage_formulas(
            ws,
            calc_rows,
            current_base_cell=f"$D${ingresos_row}",
            previous_base_cell=f"$F${ingresos_row}",
        )

    row += 1
    _write_signatures(ws, branding, row)

    widths = {"A": 8, "B": 48, "C": 6, "D": 16, "E": 9, "F": 16, "G": 9, "H": 16, "I": 9, "J": 8}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Notas
# ---------------------------------------------------------------------------
def _prepare_notas_data(*dfs: pd.DataFrame | None) -> pd.DataFrame:
    """Normalize report data for Notas: subgrupo, cuenta, tercero and valor."""
    frames = [df for df in dfs if df is not None and not df.empty]
    columns = ["SUBGRUPO_COD", "CUENTA", "CUENTA_COD", "CUENTA_NOMBRE", "TERCERO", "VALOR"]
    if not frames:
        return pd.DataFrame(columns=columns)

    combined = pd.concat(frames, ignore_index=True).copy()
    if "CUENTA" not in combined.columns or "VALOR" not in combined.columns:
        return pd.DataFrame(columns=columns)

    combined["CUENTA"] = combined["CUENTA"].fillna("").astype(str).str.strip()
    combined = combined[combined["CUENTA"] != ""].copy()
    if combined.empty:
        return pd.DataFrame(columns=columns)

    combined["VALOR"] = pd.to_numeric(combined["VALOR"], errors="coerce").fillna(0)
    combined["CUENTA_COD"] = combined["CUENTA"].str.extract(r'^(\d+)')[0].fillna("")
    combined["SUBGRUPO_COD"] = combined["CUENTA_COD"].str[:2]

    cuenta_parts = combined["CUENTA"].str.split(" - ", n=1, expand=True)
    if cuenta_parts.shape[1] > 1:
        combined["CUENTA_NOMBRE"] = cuenta_parts[1].fillna(combined["CUENTA"])
    else:
        combined["CUENTA_NOMBRE"] = combined["CUENTA"]
    combined["CUENTA_NOMBRE"] = combined["CUENTA_NOMBRE"].astype(str).str.strip()

    if "TERCERO" not in combined.columns:
        combined["TERCERO"] = ""
    combined["TERCERO"] = combined["TERCERO"].fillna("").astype(str).str.strip()

    return combined[columns].copy()


def _notas_account_totals(notes_df: pd.DataFrame, sg_code: str) -> dict[str, dict[str, Any]]:
    """Build account-level totals for one note/subgrupo."""
    if notes_df.empty:
        return {}

    sg_data = notes_df[notes_df["SUBGRUPO_COD"] == sg_code]
    if sg_data.empty:
        return {}

    grouped = (
        sg_data.groupby(["CUENTA", "CUENTA_COD", "CUENTA_NOMBRE"], as_index=False, dropna=False)
        .agg(VALOR=("VALOR", "sum"))
        .sort_values("CUENTA_COD")
    )

    return {
        str(row["CUENTA"]): {
            "codigo": str(row["CUENTA_COD"]),
            "nombre": str(row["CUENTA_NOMBRE"]),
            "valor": _money(row["VALOR"]),
        }
        for _, row in grouped.iterrows()
    }


def _notas_tercero_totals(notes_df: pd.DataFrame, sg_code: str, cuenta: str) -> dict[str, float]:
    """Build third-party totals for one account inside one note/subgrupo."""
    if notes_df.empty:
        return {}

    detail = notes_df[
        (notes_df["SUBGRUPO_COD"] == sg_code)
        & (notes_df["CUENTA"] == cuenta)
    ].copy()
    if detail.empty:
        return {}

    detail["TERCERO_LABEL"] = detail["TERCERO"].where(detail["TERCERO"] != "", "Sin tercero")
    grouped = detail.groupby("TERCERO_LABEL", as_index=False, dropna=False)["VALOR"].sum()
    return {
        str(row["TERCERO_LABEL"]): _money(row["VALOR"])
        for _, row in grouped.sort_values("TERCERO_LABEL").iterrows()
    }


def _notas_pct_variacion(valor: float, valor_anterior: float) -> float:
    """Return variation percentage using the previous period as base."""
    return (valor - valor_anterior) / abs(valor_anterior) if valor_anterior != 0 else 0


def _notas_var_formula(row: int) -> str:
    """Excel formula for value variation in the Notas sheet."""
    return f"=C{row}-D{row}"


def _notas_pct_formula(row: int) -> str:
    """Excel formula for percentage variation in the Notas sheet."""
    return f"=IF(D{row}=0,0,E{row}/ABS(D{row}))"


def _notas_sum_formula(column: str, rows: list[int]) -> str:
    """Excel formula that sums selected rows without including hidden detail twice."""
    if not rows:
        return "0"
    return "=SUM(" + ",".join(f"{column}{row}" for row in rows) + ")"


def _build_notas_sheet(
    wb,
    df_balance,
    df_er,
    branding,
    periodo_actual,
    periodo_anterior=None,
    df_balance_anterior=None,
    df_er_anterior=None,
):
    ws = wb.create_sheet("Notas")
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False)

    row = _write_header(ws, branding, "NOTAS A LOS ESTADOS FINANCIEROS",
                        f"(Expresados en pesos colombianos) — {periodo_actual}", merge_cols=6)

    current_notes = _prepare_notas_data(df_balance, df_er)
    previous_notes = _prepare_notas_data(df_balance_anterior, df_er_anterior)

    if current_notes.empty and previous_notes.empty:
        _wc(ws, row, 1, "Sin datos para generar notas.", font=FONT_BODY)
        return

    nota_num = 1
    subgrupo_codes = sorted(set(current_notes["SUBGRUPO_COD"].dropna()) | set(previous_notes["SUBGRUPO_COD"].dropna()))
    per_ant_label = periodo_anterior or "Período Anterior"

    for sg_code in subgrupo_codes:
        current_accounts = _notas_account_totals(current_notes, sg_code)
        previous_accounts = _notas_account_totals(previous_notes, sg_code)
        account_keys = sorted(
            set(current_accounts) | set(previous_accounts),
            key=lambda key: (
                current_accounts.get(key, previous_accounts.get(key, {})).get("codigo", ""),
                key,
            ),
        )

        if not account_keys:
            continue

        total_sub = sum(current_accounts.get(account, {}).get("valor", 0) for account in account_keys)
        total_sub_ant = sum(previous_accounts.get(account, {}).get("valor", 0) for account in account_keys)
        if total_sub == 0 and total_sub_ant == 0 and len(account_keys) < 2:
            continue

        niif_name = NIIF_NAMES.get(sg_code, ER_NIIF_NAMES.get(sg_code, sg_code))

        # Nota header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        _wc(ws, row, 1, f"NOTA {nota_num}    {niif_name}",
            font=FONT_SECTION, fill=FILL_TEAL_BG, border=THIN_BORDER)
        for c in range(2, 7):
            _sc(ws.cell(row=row, column=c), fill=FILL_TEAL_BG, border=THIN_BORDER)
        row += 1

        # Column headers
        for i, h in enumerate(["CTA", "Cuenta / Tercero", periodo_actual, per_ant_label, "Variación", "% Variación"], 1):
            _wc(ws, row, i, h, font=FONT_HEADER, fill=FILL_TEAL_LIGHT, border=THIN_BORDER, alignment=ALIGN_CENTER)
        row += 1

        account_rows = []
        for account in account_keys:
            current_info = current_accounts.get(account)
            previous_info = previous_accounts.get(account)
            info = current_info or previous_info or {}

            codigo = info.get("codigo", "")
            nombre = info.get("nombre", account)
            account_row = row
            account_rows.append(account_row)

            _wc(ws, row, 1, codigo, font=FONT_BODY_BOLD, border=THIN_BORDER)
            _wc(ws, row, 2, nombre, font=FONT_BODY_BOLD, border=THIN_BORDER, alignment=ALIGN_LEFT)
            _wc(ws, row, 3, "", font=FONT_BODY_BOLD, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
            _wc(ws, row, 4, "", font=FONT_BODY_BOLD, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
            _wc(ws, row, 5, "", font=FONT_BODY_BOLD, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
            _wc(ws, row, 6, "", font=FONT_BODY_BOLD, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT)
            row += 1

            current_terceros = _notas_tercero_totals(current_notes, sg_code, account)
            previous_terceros = _notas_tercero_totals(previous_notes, sg_code, account)
            tercero_keys = sorted(set(current_terceros) | set(previous_terceros))
            detail_start_row = row

            for tercero in tercero_keys:
                tercero_val = current_terceros.get(tercero, 0)
                tercero_val_ant = previous_terceros.get(tercero, 0)

                _wc(ws, row, 1, "", font=FONT_DETAIL, border=THIN_BORDER)
                _wc(ws, row, 2, f"    {tercero}", font=FONT_DETAIL, border=THIN_BORDER, alignment=ALIGN_LEFT)
                _wc(ws, row, 3, tercero_val, font=FONT_DETAIL, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
                _wc(ws, row, 4, tercero_val_ant, font=FONT_DETAIL, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
                _wc(ws, row, 5, _notas_var_formula(row), font=FONT_DETAIL, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
                _wc(ws, row, 6, _notas_pct_formula(row), font=FONT_DETAIL, border=THIN_BORDER, alignment=ALIGN_RIGHT, number_format=PCT_FMT)
                ws.row_dimensions[row].outline_level = 1
                ws.row_dimensions[row].hidden = True
                row += 1

            detail_end_row = row - 1
            if detail_start_row <= detail_end_row:
                ws.cell(row=account_row, column=3).value = f"=SUM(C{detail_start_row}:C{detail_end_row})"
                ws.cell(row=account_row, column=4).value = f"=SUM(D{detail_start_row}:D{detail_end_row})"
            else:
                ws.cell(row=account_row, column=3).value = current_info.get("valor", 0) if current_info else 0
                ws.cell(row=account_row, column=4).value = previous_info.get("valor", 0) if previous_info else 0
            ws.cell(row=account_row, column=5).value = _notas_var_formula(account_row)
            ws.cell(row=account_row, column=6).value = _notas_pct_formula(account_row)

        _wc(ws, row, 1, "", fill=FILL_TEAL_LIGHT, border=THIN_BORDER)
        _wc(ws, row, 2, "Total", font=FONT_BODY_BOLD, fill=FILL_TEAL_LIGHT, border=THIN_BORDER)
        _wc(ws, row, 3, _notas_sum_formula("C", account_rows),
            font=FONT_BODY_BOLD, fill=FILL_TEAL_LIGHT, border=THIN_BORDER,
            alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
        _wc(ws, row, 4, _notas_sum_formula("D", account_rows),
            font=FONT_BODY_BOLD, fill=FILL_TEAL_LIGHT, border=THIN_BORDER,
            alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
        _wc(ws, row, 5, _notas_var_formula(row),
            font=FONT_BODY_BOLD, fill=FILL_TEAL_LIGHT, border=THIN_BORDER,
            alignment=ALIGN_RIGHT, number_format=NUM_FMT_NEG)
        _wc(ws, row, 6, _notas_pct_formula(row),
            font=FONT_BODY_BOLD, fill=FILL_TEAL_LIGHT, border=THIN_BORDER,
            alignment=ALIGN_RIGHT, number_format=PCT_FMT)
        row += 2

        nota_num += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12


# ---------------------------------------------------------------------------
# Graficas
# ---------------------------------------------------------------------------
def _build_graficas_sheet(wb, df_balance, df_er, branding):
    ws = wb.create_sheet("Graficas")

    _wc(ws, 1, 1, branding.get("empresa", ""), font=FONT_TITLE)
    _wc(ws, 2, 1, "GRÁFICAS DE COMPOSICIÓN FINANCIERA", font=FONT_SUBTITLE)

    _build_pie(ws, df_balance, "ACTIVO", "Composición del Activo", 5, "A20", 0)
    _build_pie(ws, df_balance, "PASIVO", "Composición del Pasivo", 5, "J20", 5)
    _build_bar(ws, df_er, 5, "A38")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 15


def _build_pie(ws, df, clase, title, data_row, anchor, col_off):
    data = df[df["CLASE"] == clase]
    if data.empty:
        return

    data = data.copy()
    data["SG"] = data["CUENTA"].astype(str).str.extract(r'^(\d{2})')[0]
    grouped = data.groupby("SG", as_index=False)["VALOR"].sum()
    grouped = grouped[grouped["VALOR"] != 0]
    if grouped.empty:
        return

    grouped["LABEL"] = grouped["SG"].map(lambda x: NIIF_NAMES.get(x, x))
    source_ws = ws.parent["EF"] if "EF" in ws.parent.sheetnames else None

    def _find_ef_row(label):
        if source_ws is None:
            return None
        for source_row in range(1, source_ws.max_row + 1):
            if source_ws.cell(source_row, 2).value == label:
                return source_row
        return None

    cl, cv = 1 + col_off, 2 + col_off
    _wc(ws, data_row, cl, "Concepto", font=FONT_HEADER, fill=FILL_TEAL_LIGHT, border=THIN_BORDER)
    _wc(ws, data_row, cv, "Valor", font=FONT_HEADER, fill=FILL_TEAL_LIGHT, border=THIN_BORDER)

    for i, (_, r) in enumerate(grouped.iterrows()):
        rr = data_row + 1 + i
        label = r["LABEL"]
        source_row = _find_ef_row(label)
        value = f"=ABS('EF'!D{source_row})" if source_row else abs(r["VALOR"])
        _wc(ws, rr, cl, label, font=FONT_BODY, border=THIN_BORDER)
        _wc(ws, rr, cv, value, font=FONT_BODY, border=THIN_BORDER, number_format=NUM_FMT)

    last = data_row + len(grouped)
    pie = PieChart()
    pie.title = title
    pie.style = 10
    pie.width, pie.height = 18, 14
    labels = Reference(ws, min_col=cl, min_row=data_row + 1, max_row=last)
    vals = Reference(ws, min_col=cv, min_row=data_row, max_row=last)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, anchor)


def _build_bar(ws, df_er, data_row, anchor):
    if df_er.empty:
        return

    df_er = df_er.copy()
    df_er["SG"] = df_er["CUENTA"].astype(str).str.extract(r'^(\d{2})')[0]

    items = {}
    for sg, label in [("41", "Ingresos"), ("51", "Gastos Admin"), ("52", "Gastos Venta"),
                       ("61", "Costo Ventas"), ("53", "Gastos No Op")]:
        d = df_er[df_er["SG"] == sg]
        v = abs(d["VALOR"].sum()) if not d.empty else 0
        if v > 0:
            items[label] = v

    if not items:
        return

    source_ws = ws.parent["ER"] if "ER" in ws.parent.sheetnames else None
    source_labels = {
        "Ingresos": "Venta de producto y prestación de servicios",
        "Gastos Admin": "Total Gastos de administración",
        "Gastos Venta": "Total Gastos de Venta",
        "Costo Ventas": "MENOS:  COSTO DE VENTAS",
        "Gastos No Op": "Gastos no ordinarios",
    }

    def _find_er_row(label):
        if source_ws is None:
            return None
        for source_row in range(1, source_ws.max_row + 1):
            if source_ws.cell(source_row, 2).value == label:
                return source_row
        return None

    col = 10
    _wc(ws, data_row, col, "Concepto", font=FONT_HEADER, fill=FILL_TEAL_LIGHT, border=THIN_BORDER)
    _wc(ws, data_row, col + 1, "Valor", font=FONT_HEADER, fill=FILL_TEAL_LIGHT, border=THIN_BORDER)

    for i, (label, val) in enumerate(items.items()):
        rr = data_row + 1 + i
        source_row = _find_er_row(source_labels.get(label, ""))
        value = f"=ABS('ER'!D{source_row})" if source_row else val
        _wc(ws, rr, col, label, font=FONT_BODY, border=THIN_BORDER)
        _wc(ws, rr, col + 1, value, font=FONT_BODY, border=THIN_BORDER, number_format=NUM_FMT)

    last = data_row + len(items)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Ingresos vs Gastos vs Costos"
    bar.style = 10
    bar.width, bar.height = 18, 14
    bar.y_axis.title = "Pesos ($)"
    cats = Reference(ws, min_col=col, min_row=data_row + 1, max_row=last)
    vals = Reference(ws, min_col=col + 1, min_row=data_row, max_row=last)
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, anchor)


# ===========================================================================
# PUBLIC API
# ===========================================================================
def generate_informe(
    df_balance: pd.DataFrame,
    df_er: pd.DataFrame,
    branding: dict,
    periodo_actual: str,
    periodo_anterior: str | None = None,
    df_balance_anterior: pd.DataFrame | None = None,
    df_er_anterior: pd.DataFrame | None = None,
) -> bytes:
    """
    Genera el INFORME completo como archivo Excel (.xlsx) en memoria,
    replicando el formato del INFORME.xls profesional.
    """
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    nota_next = _build_ef_sheet(wb, df_balance, branding, periodo_actual,
                                 periodo_anterior, df_balance_anterior)
    _build_er_sheet(wb, df_er, branding, periodo_actual, nota_start=nota_next,
                    periodo_anterior=periodo_anterior, df_er_anterior=df_er_anterior)
    _build_notas_sheet(
        wb,
        df_balance,
        df_er,
        branding,
        periodo_actual,
        periodo_anterior=periodo_anterior,
        df_balance_anterior=df_balance_anterior,
        df_er_anterior=df_er_anterior,
    )
    _build_graficas_sheet(wb, df_balance, df_er, branding)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
