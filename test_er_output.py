from pathlib import Path
from io import BytesIO
from unittest.mock import MagicMock
import sys
import unittest

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parent
SIIGO_BALANCE = ROOT / "Balance de prueba por tercero vf 2.xlsx"
REFERENCE_REPORT = ROOT / "INFORME MARZO 2026 - CONTABLE 2.xlsx"


class EstadoResultadosOutputTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        sys.modules.setdefault("streamlit", MagicMock())
        from streamlit_app import ensure_optional_columns, process_dataframe, read_siigo_excel

        with SIIGO_BALANCE.open("rb") as source:
            raw_df = ensure_optional_columns(read_siigo_excel(source))

        processed = process_dataframe(
            raw_df,
            mes="MARZO",
            estado="REAL",
            anio="2026",
            centro_costos="CAPSULAS",
            desglosar_por_tercero=True,
        )
        cls.generated_balance = processed[0]
        cls.generated_er = processed[1]

        reference_er = pd.read_excel(REFERENCE_REPORT, sheet_name="BASE DE DATOS E.R.")
        cls.reference_er = reference_er[
            (reference_er["MES"].astype(str).str.upper() == "MARZO")
            & (reference_er["AÑO"] == 2026)
            & (reference_er["ESTADO"].astype(str).str.upper() == "REAL")
            & reference_er["CUENTA"].notna()
        ].copy()

        from report_generator import generate_informe, make_branding

        cls.report_workbook = openpyxl.load_workbook(
            BytesIO(
                generate_informe(
                    cls.generated_balance,
                    cls.generated_er,
                    make_branding(empresa="PANGEA NATURAL PRODUCTS S.A.S."),
                    periodo_actual="marzo de 2026",
                )
            ),
            data_only=False,
        )

    def test_estado_resultados_has_legacy_structure(self):
        expected_columns = [
            "MES",
            "ESTADO",
            "AÑO",
            "CENTRO DE COSTOS",
            "CLASE",
            "GRUPO",
            "SUBGRUPO",
            "CUENTA",
            "TERCERO",
            "VALOR",
        ]

        self.assertEqual(list(self.generated_er.columns), expected_columns)

    def test_column_g_categories_match_reference(self):
        merged = self._reference_to_generated_merge()

        missing = merged[merged["_merge"] != "both"]
        self.assertTrue(missing.empty, self._format_rows("Missing reference rows", missing))

        category_mismatch = merged[
            merged["SUBGRUPO_ref"].astype(str).str.upper()
            != merged["SUBGRUPO_gen"].astype(str).str.upper()
        ]
        self.assertTrue(
            category_mismatch.empty,
            self._format_rows("Category mismatches", category_mismatch),
        )

    def test_values_match_reference_to_the_cent(self):
        merged = self._reference_to_generated_merge()
        valor_ref = pd.to_numeric(merged["VALOR_ref"], errors="coerce").fillna(0)
        valor_gen = pd.to_numeric(merged["VALOR_gen"], errors="coerce").fillna(0)
        cents_delta = (
            valor_ref
            - valor_gen
        ).round(2)

        value_mismatch = merged[cents_delta != 0].copy()
        value_mismatch["DIFERENCIA"] = cents_delta[cents_delta != 0]

        self.assertTrue(
            value_mismatch.empty,
            self._format_rows("Value mismatches", value_mismatch),
        )

    def test_final_report_notas_are_grouped_by_four_digits(self):
        ws = self.report_workbook["Notas"]
        group_row = self._find_row(ws, "1355")
        account_row = self._find_row(ws, "13551501")

        self.assertEqual(ws.row_dimensions[group_row].outline_level, 0)
        self.assertFalse(ws.row_dimensions[group_row].hidden)
        self.assertEqual(ws.row_dimensions[account_row].outline_level, 1)
        self.assertTrue(ws.row_dimensions[account_row].hidden)
        self.assertTrue(str(ws.cell(group_row, 3).value).startswith("=SUM("))
        self.assertTrue(str(ws.cell(account_row, 3).value).startswith("=SUM("))

    def test_final_report_er_uses_four_digit_dropdowns_and_formulas(self):
        ws = self.report_workbook["ER"]
        group_row = self._find_label_row(ws, "4135 - COMERCIO AL POR MAYOR Y AL DETAL")

        self.assertEqual(ws.row_dimensions[group_row].outline_level, 1)
        self.assertTrue(ws.row_dimensions[group_row].hidden)
        self.assertTrue(str(ws.cell(group_row, 4).value).startswith("=SUM("))

    def _reference_to_generated_merge(self):
        generated = self.generated_er.copy()
        generated["CODIGO"] = generated["CUENTA"].astype(str).str.extract(r"^(\d+)")[0]
        generated["TERCERO_NORM"] = generated["TERCERO"].fillna("").astype(str).str.strip()

        reference = self.reference_er.copy()
        reference["CODIGO"] = (
            reference["CUENTA"].astype(str).str.extract(r"^(\s*\d+)")[0].str.strip()
        )
        reference["TERCERO_NORM"] = reference["TERCERO"].fillna("").astype(str).str.strip()

        generated_agg = generated.groupby(
            ["CODIGO", "TERCERO_NORM", "GRUPO", "SUBGRUPO"],
            dropna=False,
            as_index=False,
        )["VALOR"].sum()
        reference_agg = reference.groupby(
            ["CODIGO", "TERCERO_NORM", "GRUPO", "SUBGRUPO"],
            dropna=False,
            as_index=False,
        )["VALOR"].sum()

        return reference_agg.merge(
            generated_agg,
            on=["CODIGO", "TERCERO_NORM"],
            how="left",
            suffixes=("_ref", "_gen"),
            indicator=True,
        )

    @staticmethod
    def _format_rows(title, rows):
        columns = [
            "CODIGO",
            "TERCERO_NORM",
            "GRUPO_ref",
            "GRUPO_gen",
            "SUBGRUPO_ref",
            "SUBGRUPO_gen",
            "VALOR_ref",
            "VALOR_gen",
        ]
        available_columns = [column for column in columns if column in rows.columns]
        return f"{title}:\n{rows[available_columns].head(20).to_string(index=False)}"

    @staticmethod
    def _find_row(ws, code):
        for row in range(1, ws.max_row + 1):
            if str(ws.cell(row, 1).value) == code:
                return row
        raise AssertionError(f"Code {code} not found in {ws.title}")

    @staticmethod
    def _find_label_row(ws, label):
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, 2).value
            if isinstance(value, str) and label in value:
                return row
        raise AssertionError(f"Label {label} not found in {ws.title}")


if __name__ == "__main__":
    unittest.main()
